import openpyxl

wb_b = openpyxl.load_workbook('BruceBanner.xlsx')
wb_i = openpyxl.load_workbook('BruceiCIMS.xlsx')

sheet_b = wb_b.get_sheet_by_name('Sheet1')
sheet_i = wb_i.get_sheet_by_name('Sheet1')


#in case one dataset has more columns than another
def find_cutoff():
    if sheet_b.max_column == sheet_i.max_column:
        print('Column count: Match\n')
        return sheet_b.max_column
    else:
        print('Warning: # of columns differ between datasets')
        return min(sheet_b.max_column, sheet_i.max_column)


row_cutoff = sheet_b.max_row
column_cutoff = find_cutoff()+1



def compare_data():
    for item in range(1, column_cutoff):
        print(sheet_b.cell(row=row_cutoff, column=item).value)
        print(sheet_i.cell(row=row_cutoff, column=item).value)


        
compare_data()
